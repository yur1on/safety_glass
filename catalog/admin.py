import csv
import re
from io import BytesIO, TextIOWrapper

from django.contrib import admin, messages
from django.db import transaction
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from django.utils import timezone

from openpyxl import Workbook, load_workbook

from .models import GlassGroup, Glass, GlassAlias


# ---------------------------
# Utilities
# ---------------------------
def _normalize(s: str) -> str:
    return " ".join((s or "").strip().lower().split())


def _split_aliases(raw: str) -> list[str]:
    raw = raw or ""
    parts: list[str] = []
    for chunk in raw.replace("|", ";").replace(",", ";").split(";"):
        v = chunk.strip()
        if v:
            parts.append(v)
    return parts


def sync_glass_aliases(glass: Glass) -> None:
    parts = _split_aliases(glass.aliases_text or "")

    desired_norm_to_original: dict[str, str] = {}
    for p in parts:
        n = _normalize(p)
        if n:
            desired_norm_to_original[n] = p

    existing = list(GlassAlias.objects.filter(glass=glass))
    existing_norm_to_obj = {_normalize(a.alias): a for a in existing}

    for norm, obj in existing_norm_to_obj.items():
        if norm not in desired_norm_to_original:
            obj.delete()

    for norm, original in desired_norm_to_original.items():
        if norm not in existing_norm_to_obj:
            GlassAlias.objects.create(glass=glass, alias=original)


def _append_brand_list(raw_brand_cell: str) -> str:
    raw = (raw_brand_cell or "").strip()
    if not raw:
        return ""
    parts = [p.strip() for p in raw.split(",") if p.strip()]
    seen = set()
    out: list[str] = []
    for p in parts:
        b = p.upper()
        nb = _normalize(b)
        if nb and nb not in seen:
            seen.add(nb)
            out.append(b)
    return ", ".join(out)


def _parse_models_cell(raw_models: str) -> list[str]:
    raw = (raw_models or "").strip()
    if not raw:
        return []
    models = [x.strip() for x in re.split(r"[\/\n]+", raw) if x and x.strip()]
    seen = set()
    uniq: list[str] = []
    for m in models:
        nm = _normalize(m)
        if nm and nm not in seen:
            seen.add(nm)
            uniq.append(m)
    return uniq


def _find_columns_by_header(header_row: list[str | None]) -> tuple[int, int, int]:
    """
    Находит индексы колонок: id, бренд, взаимозаменяемое стекло
    """
    id_idx = -1
    brand_idx = -1
    models_idx = -1

    for i, v in enumerate(header_row):
        hv = _normalize(str(v or ""))
        if hv == "id":
            id_idx = i
        if hv == "бренд":
            brand_idx = i
        if hv.startswith("взаимозаменяемое стекло"):
            models_idx = i

    if id_idx == -1 or brand_idx == -1 or models_idx == -1:
        raise ValueError(
            "Не найдены колонки 'id', 'бренд' и/или 'взаимозаменяемое стекло' в первой строке."
        )

    return id_idx, brand_idx, models_idx


# ---------------------------
# Admin Inlines
# ---------------------------
class GlassInline(admin.TabularInline):
    model = Glass
    extra = 5
    fields = ("name", "is_active", "aliases_text")
    readonly_fields = ()
    show_change_link = True
    ordering = ("-is_active", "name")



# ---------------------------
# Admin: GlassGroup
# ---------------------------
@admin.register(GlassGroup)
class GlassGroupAdmin(admin.ModelAdmin):
    list_display = ("name", "external_id", "brands", "created_at")
    search_fields = ("name", "external_id", "brands", "description")
    ordering = ("name",)
    fields = ("external_id", "name", "brands", "description")
    inlines = (GlassInline,)

    actions = ("export_groups_csv", "export_groups_xlsx")
    change_list_template = "admin/catalog/glassgroup/change_list.html"

    @transaction.atomic
    def save_formset(self, request, form, formset, change):
        instances = formset.save()
        for obj in instances:
            if isinstance(obj, Glass):
                sync_glass_aliases(obj)

    # ---------- export CSV ----------
    @admin.action(description="Экспорт выбранных групп в CSV")
    def export_groups_csv(self, request: HttpRequest, queryset):
        queryset = queryset.prefetch_related("glasses").order_by("name")
        ts = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"glass_groups_{ts}.csv"

        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response.write("\ufeff")

        writer = csv.writer(response)
        writer.writerow(["excel_id", "group_name", "group_brands", "group_description", "glass_name", "glass_aliases_text"])

        for group in queryset:
            glasses = list(group.glasses.all().order_by("name"))
            if not glasses:
                writer.writerow([group.external_id, group.name, group.brands or "", group.description or "", "", ""])
                continue
            for glass in glasses:
                writer.writerow([group.external_id, group.name, group.brands or "", group.description or "", glass.name, glass.aliases_text or ""])

        return response

    # ---------- export XLSX ----------
    @admin.action(description="Экспорт выбранных групп в Excel (XLSX)")
    def export_groups_xlsx(self, request: HttpRequest, queryset):
        queryset = queryset.prefetch_related("glasses").order_by("name")

        wb = Workbook()
        ws = wb.active
        ws.title = "Glass Groups"
        ws.append(["excel_id", "group_name", "group_brands", "group_description", "glass_name", "glass_aliases_text"])

        for group in queryset:
            glasses = list(group.glasses.all().order_by("name"))
            if not glasses:
                ws.append([group.external_id, group.name, group.brands or "", group.description or "", "", ""])
                continue
            for glass in glasses:
                ws.append([group.external_id, group.name, group.brands or "", group.description or "", glass.name, glass.aliases_text or ""])

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        ts = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"glass_groups_{ts}.xlsx"

        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    # ---------- custom URLs ----------
    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path("import-csv/", self.admin_site.admin_view(self.import_csv_view), name="catalog_glassgroup_import_csv"),
            path("import-xlsx/", self.admin_site.admin_view(self.import_xlsx_view), name="catalog_glassgroup_import_xlsx"),
        ]
        return custom + urls

    # ---------- import CSV (как раньше) ----------
    @transaction.atomic
    def import_csv_view(self, request: HttpRequest):
        if request.method == "POST":
            uploaded = request.FILES.get("csv_file")
            if not uploaded:
                messages.error(request, "Файл не выбран.")
                return redirect("..")

            clear = bool(request.POST.get("clear_before_import"))
            if clear:
                GlassAlias.objects.all().delete()
                Glass.objects.all().delete()
                GlassGroup.objects.all().delete()
                messages.warning(request, "Данные очищены перед импортом.")

            try:
                wrapper = TextIOWrapper(uploaded.file, encoding="utf-8-sig", newline="")
                reader = csv.DictReader(wrapper)

                required = {"group_name", "group_brands", "group_description", "glass_name", "glass_aliases_text"}
                if not required.issubset(set(reader.fieldnames or [])):
                    messages.error(request, "Неверные заголовки CSV.")
                    return redirect("..")

                for row in reader:
                    group_name = (row.get("group_name") or "").strip()
                    if not group_name:
                        continue

                    group_brands = (row.get("group_brands") or "").strip()
                    group_description = (row.get("group_description") or "").strip()
                    glass_name = (row.get("glass_name") or "").strip()
                    glass_aliases_text = (row.get("glass_aliases_text") or "").strip()

                    group, _ = GlassGroup.objects.get_or_create(name=group_name)
                    if group_brands:
                        group.brands = group_brands
                    if group_description:
                        group.description = group_description
                    group.save()

                    if not glass_name:
                        continue

                    glass, _ = Glass.objects.get_or_create(group=group, name=glass_name)
                    glass.aliases_text = glass_aliases_text
                    glass.save()
                    sync_glass_aliases(glass)

                messages.success(request, "Импорт CSV завершён.")
                return redirect("..")

            except Exception as e:
                messages.error(request, f"Ошибка импорта CSV: {e}")
                return redirect("..")

        context = {**self.admin_site.each_context(request), "title": "Импорт групп стёкол из CSV"}
        return render(request, "admin/catalog/glassgroup/import_csv.html", context)

    # ---------- import XLSX (3 колонки: id/бренд/взаимозаменяемое стекло) ----------

    @transaction.atomic
    def import_xlsx_view(self, request: HttpRequest):
        if request.method == "POST":
            uploaded = request.FILES.get("xlsx_file")
            if not uploaded:
                messages.error(request, "Файл не выбран.")
                return redirect("..")

            # вместо удаления — деактивация
            deactivate_missing_groups = bool(request.POST.get("deactivate_missing_groups"))
            deactivate_missing_models = bool(request.POST.get("deactivate_missing_models"))

            try:
                wb = load_workbook(uploaded, data_only=True)
                ws = wb[wb.sheetnames[0]]

                header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
                id_idx, brand_idx, models_idx = _find_columns_by_header(header)

                created_groups = 0
                updated_groups = 0
                created_models = 0
                reactivated_groups = 0
                reactivated_models = 0
                deactivated_groups = 0
                deactivated_models = 0

                excel_group_ids: set[str] = set()

                # --- импорт/обновление групп из файла ---
                for r in range(2, ws.max_row + 1):
                    excel_id = ws.cell(r, id_idx + 1).value
                    brand_cell = ws.cell(r, brand_idx + 1).value
                    models_cell = ws.cell(r, models_idx + 1).value

                    excel_id = "" if excel_id is None else str(excel_id).strip()
                    if not excel_id:
                        continue

                    brands = _append_brand_list("" if brand_cell is None else str(brand_cell))
                    models = _parse_models_cell("" if models_cell is None else str(models_cell))
                    if not models:
                        continue

                    excel_group_ids.add(excel_id)

                    group_name = f"{models[0]} ({excel_id})"

                    group, created = GlassGroup.objects.get_or_create(
                        external_id=excel_id,
                        defaults={"name": group_name, "brands": brands, "is_active": True},
                    )

                    if created:
                        created_groups += 1
                    else:
                        changed = False

                        # ре-активация группы (если ранее выключили)
                        if hasattr(group, "is_active") and group.is_active is False:
                            group.is_active = True
                            changed = True
                            reactivated_groups += 1

                        if group.name != group_name:
                            group.name = group_name
                            changed = True
                        if brands and (group.brands or "") != brands:
                            group.brands = brands
                            changed = True

                        if changed:
                            group.save()
                            updated_groups += 1

                    # --- синхронизация моделей внутри группы ---
                    desired_norms = {_normalize(m) for m in models}

                    existing = list(group.glasses.all())
                    existing_by_norm = {_normalize(g.name): g for g in existing}

                    # add / reactivate
                    for m in models:
                        nm = _normalize(m)
                        if nm in existing_by_norm:
                            obj = existing_by_norm[nm]
                            if hasattr(obj, "is_active") and obj.is_active is False:
                                obj.is_active = True
                                obj.save(update_fields=["is_active"])
                                reactivated_models += 1
                            continue

                        Glass.objects.create(group=group, name=m, is_active=True)
                        created_models += 1

                    # deactivate missing models (optional)
                    if deactivate_missing_models:
                        for nm, obj in existing_by_norm.items():
                            if nm not in desired_norms:
                                if hasattr(obj, "is_active") and obj.is_active is True:
                                    obj.is_active = False
                                    obj.save(update_fields=["is_active"])
                                    deactivated_models += 1

                # --- деактивация групп, которых больше нет в Excel ---
                if deactivate_missing_groups:
                    qs = GlassGroup.objects.exclude(external_id__in=excel_group_ids)
                    for g in qs:
                        if hasattr(g, "is_active") and g.is_active is True:
                            g.is_active = False
                            g.save(update_fields=["is_active"])
                            deactivated_groups += 1

                messages.success(
                    request,
                    "Импорт XLSX завершён. "
                    f"Группы: +{created_groups}, обновлено {updated_groups}, "
                    f"реактивировано {reactivated_groups}, деактивировано {deactivated_groups}. "
                    f"Модели: +{created_models}, реактивировано {reactivated_models}, деактивировано {deactivated_models}."
                )
                return redirect("..")

            except Exception as e:
                messages.error(request, f"Ошибка импорта XLSX: {e}")
                return redirect("..")

        context = {**self.admin_site.each_context(request), "title": "Импорт из Excel (ваш формат: 3 колонки)"}
        return render(request, "admin/catalog/glassgroup/import_xlsx.html", context)


@admin.register(Glass)
class GlassAdmin(admin.ModelAdmin):
    list_display = ("name", "group")
    search_fields = ("name", "group__name")
    autocomplete_fields = ("group",)
    ordering = ("group__name", "name")


@admin.register(GlassAlias)
class GlassAliasAdmin(admin.ModelAdmin):
    list_display = ("alias", "glass", "get_group")
    search_fields = ("alias", "normalized_alias", "glass__name", "glass__group__name")
    autocomplete_fields = ("glass",)
    ordering = ("alias",)

    @admin.display(description="Группа")
    def get_group(self, obj):
        return obj.glass.group


admin.site.site_header = "Администрирование базы стёкол"
admin.site.site_title = "Админка стёкол"
admin.site.index_title = "Управление справочниками"
